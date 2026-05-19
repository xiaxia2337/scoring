import sys
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
import sqlite3
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import tempfile

app = Flask(__name__)
app.secret_key = 'scoring_system_secret_key'

if getattr(sys, 'frozen', False):
    DATABASE = os.path.join(os.path.dirname(sys.executable), 'scoring_web.db')
else:
    DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scoring_web.db')

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS competitions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            num_groups INTEGER NOT NULL,
            num_judges INTEGER NOT NULL,
            num_players INTEGER NOT NULL,
            judge_weights TEXT DEFAULT "33,33,34",
            created_at TEXT NOT NULL
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS custom_names (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competition_id INTEGER NOT NULL,
            name_type TEXT NOT NULL,
            item_num INTEGER NOT NULL,
            group_num INTEGER,
            custom_name TEXT NOT NULL,
            FOREIGN KEY (competition_id) REFERENCES competitions(id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competition_id INTEGER NOT NULL,
            group_num INTEGER NOT NULL,
            player_num INTEGER NOT NULL,
            judge_num INTEGER NOT NULL,
            item_scores TEXT NOT NULL,
            total_score REAL NOT NULL,
            scored_at TEXT NOT NULL,
            FOREIGN KEY (competition_id) REFERENCES competitions(id)
        )
    ''')
    
    conn.commit()
    conn.close()

@app.route('/')
def index():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM competitions ORDER BY created_at DESC')
    competitions = cursor.fetchall()
    conn.close()
    return render_template('index.html', competitions=competitions)

@app.route('/create', methods=['GET', 'POST'])
def create_competition():
    if request.method == 'POST':
        name = request.form.get('name', '未命名比赛')
        num_groups = int(request.form.get('num_groups', 3))
        num_judges = int(request.form.get('num_judges', 3))
        num_players = int(request.form.get('num_players', 10))
        
        judge_weights = []
        for i in range(1, num_judges + 1):
            weight = float(request.form.get(f'judge_weight_{i}', round(100 / num_judges, 2)))
            judge_weights.append(weight)
        judge_weights_str = ','.join(map(str, judge_weights))
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO competitions (name, num_groups, num_judges, num_players, judge_weights, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (name, num_groups, num_judges, num_players, judge_weights_str, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        comp_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return redirect(url_for('setup_competition', comp_id=comp_id))
    
    return render_template('create.html')

@app.route('/setup/<int:comp_id>', methods=['GET', 'POST'])
def setup_competition(comp_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM competitions WHERE id = ?', (comp_id,))
    comp = cursor.fetchone()
    
    if request.method == 'POST':
        name_type = request.form.get('name_type')
        if name_type == 'group':
            for i in range(1, comp['num_groups'] + 1):
                name = request.form.get(f'group_{i}', f'第{i}组')
                cursor.execute('INSERT INTO custom_names (competition_id, name_type, item_num, custom_name) VALUES (?, ?, ?, ?)',
                            (comp_id, 'group', i, name))
        elif name_type == 'judge':
            for i in range(1, comp['num_judges'] + 1):
                name = request.form.get(f'judge_{i}', f'评委{i}')
                cursor.execute('INSERT INTO custom_names (competition_id, name_type, item_num, custom_name) VALUES (?, ?, ?, ?)',
                            (comp_id, 'judge', i, name))
        elif name_type == 'player':
            for g in range(1, comp['num_groups'] + 1):
                for i in range(1, comp['num_players'] + 1):
                    name = request.form.get(f'player_{g}_{i}', f'选手{i}')
                    cursor.execute('INSERT INTO custom_names (competition_id, name_type, item_num, group_num, custom_name) VALUES (?, ?, ?, ?, ?)',
                                (comp_id, f'player_{g}', i, g, name))
        elif name_type == 'item':
            for i in range(1, 21):
                name = request.form.get(f'item_{i}', f'项{i}')
                cursor.execute('INSERT INTO custom_names (competition_id, name_type, item_num, custom_name) VALUES (?, ?, ?, ?)',
                            (comp_id, 'item', i, name))
        
        conn.commit()
        return jsonify({'success': True})
    
    conn.close()
    return render_template('setup.html', comp=comp)

@app.route('/import_excel/<int:comp_id>', methods=['POST'])
def import_excel(comp_id):
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'})
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': '请上传Excel文件(.xlsx)'})
    
    try:
        wb = load_workbook(file)
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT num_groups, num_judges, num_players FROM competitions WHERE id = ?', (comp_id,))
        comp = cursor.fetchone()
        
        if '组别' in wb.sheetnames:
            ws = wb['组别']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    group_num = int(row[0])
                    if group_num <= comp['num_groups']:
                        cursor.execute('INSERT INTO custom_names (competition_id, name_type, item_num, custom_name) VALUES (?, ?, ?, ?)',
                                    (comp_id, 'group', group_num, str(row[1])))
        
        if '评委' in wb.sheetnames:
            ws = wb['评委']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    judge_num = int(row[0])
                    if judge_num <= comp['num_judges']:
                        cursor.execute('INSERT INTO custom_names (competition_id, name_type, item_num, custom_name) VALUES (?, ?, ?, ?)',
                                    (comp_id, 'judge', judge_num, str(row[1])))
        
        if '选手' in wb.sheetnames:
            ws = wb['选手']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2]:
                    group_num = int(row[0])
                    player_num = int(row[1])
                    if group_num <= comp['num_groups'] and player_num <= comp['num_players']:
                        cursor.execute('INSERT INTO custom_names (competition_id, name_type, item_num, group_num, custom_name) VALUES (?, ?, ?, ?, ?)',
                                    (comp_id, f'player_{group_num}', player_num, group_num, str(row[2])))
        
        if '评分项' in wb.sheetnames:
            ws = wb['评分项']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    item_num = int(row[0])
                    if item_num <= 20:
                        cursor.execute('INSERT INTO custom_names (competition_id, name_type, item_num, custom_name) VALUES (?, ?, ?, ?)',
                                    (comp_id, 'item', item_num, str(row[1])))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '导入成功'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})

@app.route('/judge_login/<int:comp_id>', methods=['GET', 'POST'])
def judge_login(comp_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM competitions WHERE id = ?', (comp_id,))
    comp = cursor.fetchone()
    
    cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "group" ORDER BY item_num', (comp_id,))
    groups = {row['item_num']: row['custom_name'] for row in cursor.fetchall()}
    
    cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "judge" ORDER BY item_num', (comp_id,))
    judges = {row['item_num']: row['custom_name'] for row in cursor.fetchall()}
    
    conn.close()
    
    if request.method == 'POST':
        group_num = int(request.form['group_num'])
        judge_num = int(request.form['judge_num'])
        
        session['comp_id'] = comp_id
        session['group_num'] = group_num
        session['judge_num'] = judge_num
        
        return redirect(url_for('scoring'))
    
    return render_template('judge_login.html', comp=comp, groups=groups, judges=judges)

@app.route('/scoring', methods=['GET', 'POST'])
def scoring():
    if 'comp_id' not in session:
        return redirect(url_for('index'))
    
    comp_id = session['comp_id']
    group_num = session['group_num']
    judge_num = session['judge_num']
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM competitions WHERE id = ?', (comp_id,))
    comp = cursor.fetchone()
    
    if not comp:
        conn.close()
        return redirect(url_for('index'))
    
    cursor.execute('SELECT custom_name FROM custom_names WHERE competition_id = ? AND name_type = "group" AND item_num = ?', (comp_id, group_num))
    group_result = cursor.fetchone()
    group_name = group_result['custom_name'] if group_result else f'第{group_num}组'
    
    cursor.execute('SELECT custom_name FROM custom_names WHERE competition_id = ? AND name_type = "judge" AND item_num = ?', (comp_id, judge_num))
    judge_result = cursor.fetchone()
    judge_name = judge_result['custom_name'] if judge_result else f'评委{judge_num}'
    
    cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "player_" || ? ORDER BY item_num', (comp_id, group_num))
    players = {row['item_num']: row['custom_name'] for row in cursor.fetchall()}
    if not players:
        for i in range(1, comp['num_players'] + 1):
            players[i] = f'选手{i}'
    
    cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "item" ORDER BY item_num', (comp_id,))
    items = {row['item_num']: row['custom_name'] for row in cursor.fetchall()}
    if not items:
        for i in range(1, 21):
            items[i] = f'项{i}'
    
    conn.close()
    
    if request.method == 'POST':
        player_num = int(request.form['player_num'])
        item_scores = []
        for i in range(1, 21):
            item_scores.append(int(request.form.get(f'score_{i}', 3)))
        
        total_score = sum(item_scores)
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM scores WHERE competition_id = ? AND group_num = ? AND player_num = ? AND judge_num = ?',
                    (comp_id, group_num, player_num, judge_num))
        existing = cursor.fetchone()
        
        if existing:
            cursor.execute('''
                UPDATE scores SET item_scores = ?, total_score = ?, scored_at = ?
                WHERE id = ?
            ''', (str(item_scores), total_score, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), existing['id']))
        else:
            cursor.execute('''
                INSERT INTO scores (competition_id, group_num, player_num, judge_num, item_scores, total_score, scored_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (comp_id, group_num, player_num, judge_num, str(item_scores), total_score, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'total_score': total_score})
    
    return render_template('scoring.html', comp=comp, group_name=group_name, judge_name=judge_name, 
                        group_num=group_num, judge_num=judge_num, players=players, items=items)

@app.route('/statistics/<int:comp_id>')
def statistics(comp_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM competitions WHERE id = ?', (comp_id,))
    comp = cursor.fetchone()
    
    cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "group" ORDER BY item_num', (comp_id,))
    groups = {row['item_num']: row['custom_name'] for row in cursor.fetchall()}
    
    cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "judge" ORDER BY item_num', (comp_id,))
    judges = {row['item_num']: row['custom_name'] for row in cursor.fetchall()}
    
    cursor.execute('SELECT * FROM custom_names WHERE competition_id = ? AND name_type LIKE "player_%"', (comp_id,))
    players = {}
    for row in cursor.fetchall():
        g = row['group_num']
        if g not in players:
            players[g] = {}
        players[g][row['item_num']] = row['custom_name']
    
    cursor.execute('''
        SELECT group_num, player_num, judge_num, total_score, scored_at
        FROM scores WHERE competition_id = ? ORDER BY group_num, player_num, judge_num
    ''', (comp_id,))
    detail_data = []
    for row in cursor.fetchall():
        detail_data.append({
            'group_name': groups.get(row['group_num'], f'第{row["group_num"]}组'),
            'player_name': players.get(row['group_num'], {}).get(row['player_num'], f'选手{row["player_num"]}'),
            'judge_name': judges.get(row['judge_num'], f'评委{row["judge_num"]}'),
            'total_score': row['total_score'],
            'scored_at': row['scored_at']
        })
    
    cursor.execute('''
        SELECT group_num, player_num, judge_num, total_score
        FROM scores WHERE competition_id = ? ORDER BY group_num, player_num, judge_num
    ''', (comp_id,))
    
    if 'judge_weights' in comp and comp['judge_weights'] and comp['judge_weights'].strip():
        judge_weights = [float(w) for w in comp['judge_weights'].split(',')]
    else:
        judge_weights = [100.0 / comp['num_judges']] * comp['num_judges']
    
    player_scores = {}
    for row in cursor.fetchall():
        key = (row['group_num'], row['player_num'])
        if key not in player_scores:
            player_scores[key] = []
        player_scores[key].append((row['judge_num'], row['total_score']))
    
    summary_data = []
    for (group_num, player_num), scores in player_scores.items():
        weighted_score = 0
        judge_count = len(scores)
        for judge_num, score in scores:
            if judge_num <= len(judge_weights):
                weight = judge_weights[judge_num - 1] / 100
                weighted_score += score * weight
        
        summary_data.append({
            'group_name': groups.get(group_num, f'第{group_num}组'),
            'player_name': players.get(group_num, {}).get(player_num, f'选手{player_num}'),
            'avg_score': round(weighted_score, 2),
            'judge_count': judge_count
        })
    
    conn.close()
    
    return render_template('statistics.html', comp=comp, detail_data=detail_data, summary_data=summary_data)

@app.route('/export_excel/<int:comp_id>')
def export_excel(comp_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM competitions WHERE id = ?', (comp_id,))
    comp = cursor.fetchone()
    
    cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "group" ORDER BY item_num', (comp_id,))
    groups = {row['item_num']: row['custom_name'] for row in cursor.fetchall()}
    
    cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "judge" ORDER BY item_num', (comp_id,))
    judges = {row['item_num']: row['custom_name'] for row in cursor.fetchall()}
    
    cursor.execute('SELECT * FROM custom_names WHERE competition_id = ? AND name_type LIKE "player_%"', (comp_id,))
    players = {}
    for row in cursor.fetchall():
        g = row['group_num']
        if g not in players:
            players[g] = {}
        players[g][row['item_num']] = row['custom_name']
    
    wb = Workbook()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center")
    
    ws1 = wb.active
    ws1.title = "详细表"
    ws1.append(["组别", "选手", "评委", "总分"])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    cursor.execute('SELECT group_num, player_num, judge_num, total_score FROM scores WHERE competition_id = ? ORDER BY group_num, player_num, judge_num', (comp_id,))
    for row in cursor.fetchall():
        group_name = groups.get(row['group_num'], f'第{row["group_num"]}组')
        player_name = players.get(row['group_num'], {}).get(row['player_num'], f'选手{row["player_num"]}')
        judge_name = judges.get(row['judge_num'], f'评委{row["judge_num"]}')
        ws1.append([group_name, player_name, judge_name, round(row['total_score'], 2)])
    
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
    
    ws2 = wb.create_sheet(title="汇总表")
    ws2.append(["组别", "选手", "平均分", f"评分评委数/{comp['num_judges']}"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    cursor.execute('SELECT group_num, player_num, judge_num, total_score FROM scores WHERE competition_id = ? ORDER BY group_num, player_num, judge_num', (comp_id,))
    
    judge_weights_str = comp.get('judge_weights', '')
    if judge_weights_str and judge_weights_str.strip():
        judge_weights = [float(w) for w in judge_weights_str.split(',')]
    else:
        judge_weights = [100.0 / comp['num_judges']] * comp['num_judges']
    
    player_scores = {}
    for row in cursor.fetchall():
        key = (row['group_num'], row['player_num'])
        if key not in player_scores:
            player_scores[key] = []
        player_scores[key].append((row['judge_num'], row['total_score']))
    
    for (group_num, player_num), scores in player_scores.items():
        weighted_score = 0
        judge_count = len(scores)
        for judge_num, score in scores:
            if judge_num <= len(judge_weights):
                weight = judge_weights[judge_num - 1] / 100
                weighted_score += score * weight
        
        group_name = groups.get(group_num, f'第{group_num}组')
        player_name = players.get(group_num, {}).get(player_num, f'选手{player_num}')
        ws2.append([group_name, player_name, round(weighted_score, 2), f"{judge_count}/{comp['num_judges']}"])
    
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
    
    conn.close()
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(temp_file.name, as_attachment=True, download_name=f'{comp["name"]}_评分统计.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/import_scores/<int:comp_id>', methods=['POST'])
def import_scores(comp_id):
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'})
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': '请上传Excel文件(.xlsx)'})
    
    try:
        wb = load_workbook(file)
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT num_groups, num_judges, num_players FROM competitions WHERE id = ?', (comp_id,))
        comp = cursor.fetchone()
        
        cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "group" ORDER BY item_num', (comp_id,))
        groups = {row['custom_name']: row['item_num'] for row in cursor.fetchall()}
        
        cursor.execute('SELECT item_num, custom_name FROM custom_names WHERE competition_id = ? AND name_type = "judge" ORDER BY item_num', (comp_id,))
        judges = {row['custom_name']: row['item_num'] for row in cursor.fetchall()}
        
        cursor.execute('SELECT * FROM custom_names WHERE competition_id = ? AND name_type LIKE "player_%"', (comp_id,))
        players = {}
        for row in cursor.fetchall():
            g = row['group_num']
            if g not in players:
                players[g] = {}
            players[g][row['custom_name']] = row['item_num']
        
        if '详细表' in wb.sheetnames:
            ws = wb['详细表']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[2] and row[3]:
                    group_name = str(row[0])
                    player_name = str(row[1])
                    judge_name = str(row[2])
                    total_score = float(row[3])
                    
                    group_num = groups.get(group_name)
                    if not group_num:
                        continue
                    
                    judge_num = judges.get(judge_name)
                    if not judge_num:
                        continue
                    
                    player_num = players.get(group_num, {}).get(player_name)
                    if not player_num:
                        continue
                    
                    item_scores = [3] * 20
                    cursor.execute('SELECT id FROM scores WHERE competition_id = ? AND group_num = ? AND player_num = ? AND judge_num = ?',
                                (comp_id, group_num, player_num, judge_num))
                    existing = cursor.fetchone()
                    
                    if existing:
                        cursor.execute('UPDATE scores SET total_score = ?, scored_at = ? WHERE id = ?',
                                    (total_score, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), existing['id']))
                    else:
                        cursor.execute('INSERT INTO scores (competition_id, group_num, player_num, judge_num, item_scores, total_score, scored_at) VALUES (?, ?, ?, ?, ?, ?, ?)',
                                    (comp_id, group_num, player_num, judge_num, str(item_scores), total_score, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '导入成功'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})

@app.route('/api/check_score/<int:comp_id>/<int:group_num>/<int:player_num>/<int:judge_num>')
def check_score(comp_id, group_num, player_num, judge_num):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT item_scores FROM scores WHERE competition_id = ? AND group_num = ? AND player_num = ? AND judge_num = ?',
                (comp_id, group_num, player_num, judge_num))
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return jsonify({'exists': True, 'scores': eval(result['item_scores'])})
    return jsonify({'exists': False})

@app.route('/delete/<int:comp_id>')
def delete_competition(comp_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM scores WHERE competition_id = ?', (comp_id,))
    cursor.execute('DELETE FROM custom_names WHERE competition_id = ?', (comp_id,))
    cursor.execute('DELETE FROM competitions WHERE id = ?', (comp_id,))
    conn.commit()
    conn.close()
    
    return redirect(url_for('index'))

init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)