from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import tempfile

app = Flask(__name__)
app.secret_key = 'scoring_system_secret_key'

import sqlite3

try:
    import psycopg2
    import psycopg2.extras
    HAS_POSTGRES = True
except ImportError:
    HAS_POSTGRES = False

def get_db():
    try:
        if HAS_POSTGRES:
            host = os.environ.get('SUPABASE_HOST')
            database = os.environ.get('SUPABASE_DB')
            user = os.environ.get('SUPABASE_USER')
            password = os.environ.get('SUPABASE_PASSWORD')
            port = os.environ.get('SUPABASE_PORT', '5432')
            
            if all([host, database, user, password]):
                print("Using PostgreSQL database")
                conn = psycopg2.connect(
                    host=host,
                    database=database,
                    user=user,
                    password=password,
                    port=port
                )
                return conn
        
        print("Using SQLite database")
        DATABASE = os.environ.get('DATABASE_URL', os.path.join(tempfile.gettempdir(), 'scoring_web.db'))
        conn = sqlite3.connect(DATABASE)
        conn.row_factory = sqlite3.Row
        return conn
    except Exception as e:
        print(f"Database connection error: {str(e)}. Using SQLite fallback.")
        DATABASE = os.path.join(tempfile.gettempdir(), 'scoring_web.db')
        conn = sqlite3.connect(DATABASE)
        conn.row_factory = sqlite3.Row
        return conn

def get_db_type():
    if HAS_POSTGRES:
        host = os.environ.get('SUPABASE_HOST')
        database = os.environ.get('SUPABASE_DB')
        user = os.environ.get('SUPABASE_USER')
        password = os.environ.get('SUPABASE_PASSWORD')
        if all([host, database, user, password]):
            return 'postgres'
    return 'sqlite'

def init_db():
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS contests (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    date TEXT NOT NULL,
                    status INTEGER DEFAULT 0,
                    judge_password TEXT NOT NULL,
                    num_groups INTEGER DEFAULT 1,
                    num_judges INTEGER DEFAULT 3,
                    num_players INTEGER DEFAULT 10,
                    created_at TEXT NOT NULL
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS contestants (
                    id SERIAL PRIMARY KEY,
                    contest_id INTEGER NOT NULL REFERENCES contests(id),
                    name TEXT NOT NULL,
                    team TEXT NOT NULL,
                    number INTEGER NOT NULL
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS scores (
                    id SERIAL PRIMARY KEY,
                    contest_id INTEGER NOT NULL REFERENCES contests(id),
                    contestant_id INTEGER NOT NULL REFERENCES contestants(id),
                    judge_name TEXT NOT NULL,
                    score1 REAL,
                    score2 REAL,
                    score3 REAL,
                    score4 REAL,
                    score5 REAL,
                    total_score REAL,
                    created_at TEXT NOT NULL
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS judges (
                    id SERIAL PRIMARY KEY,
                    contest_id INTEGER NOT NULL REFERENCES contests(id),
                    name TEXT NOT NULL
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS groups (
                    id SERIAL PRIMARY KEY,
                    contest_id INTEGER NOT NULL REFERENCES contests(id),
                    name TEXT NOT NULL
                )
            ''')
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS contests (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    date TEXT NOT NULL,
                    status INTEGER DEFAULT 0,
                    judge_password TEXT NOT NULL,
                    num_groups INTEGER DEFAULT 1,
                    num_judges INTEGER DEFAULT 3,
                    num_players INTEGER DEFAULT 10,
                    created_at TEXT NOT NULL
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS contestants (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    contest_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    team TEXT NOT NULL,
                    number INTEGER NOT NULL,
                    FOREIGN KEY (contest_id) REFERENCES contests(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS scores (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    contest_id INTEGER NOT NULL,
                    contestant_id INTEGER NOT NULL,
                    judge_name TEXT NOT NULL,
                    score1 REAL,
                    score2 REAL,
                    score3 REAL,
                    score4 REAL,
                    score5 REAL,
                    total_score REAL,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (contest_id) REFERENCES contests(id),
                    FOREIGN KEY (contestant_id) REFERENCES contestants(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS judges (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    contest_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    FOREIGN KEY (contest_id) REFERENCES contests(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS groups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    contest_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    FOREIGN KEY (contest_id) REFERENCES contests(id)
                )
            ''')
        
        conn.commit()
        conn.close()
        print("Database initialized successfully")
    except Exception as e:
        print(f"Database initialization error: {str(e)}")

init_db()

@app.route('/')
def index():
    try:
        conn = get_db()
        if get_db_type() == 'postgres':
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        else:
            cursor = conn.cursor()
        cursor.execute('SELECT * FROM contests ORDER BY created_at DESC')
        contests = cursor.fetchall()
        conn.close()
        return render_template('index.html', competitions=contests)
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/create', methods=['GET', 'POST'])
def create():
    try:
        if request.method == 'POST':
            print(f"Form data: {dict(request.form)}")
            name = request.form.get('name', '未命名比赛')
            num_groups = int(request.form.get('num_groups', '1'))
            num_judges = int(request.form.get('num_judges', '3'))
            num_players = int(request.form.get('num_players', '10'))
            
            # 收集评委权重
            judge_weights = []
            for i in range(1, num_judges + 1):
                weight = request.form.get(f'judge_weight_{i}', str(100 / num_judges))
                judge_weights.append(weight)
            
            conn = get_db()
            cursor = conn.cursor()
            
            if get_db_type() == 'postgres':
                cursor.execute('''
                    INSERT INTO contests (name, date, status, judge_password, num_groups, num_judges, num_players, created_at)
                    VALUES (%s, %s, 0, %s, %s, %s, %s, %s) RETURNING id
                ''', (name, datetime.now().isoformat(), '123456', num_groups, num_judges, num_players, datetime.now().isoformat()))
                contest_id = cursor.fetchone()[0]
            else:
                cursor.execute('''
                    INSERT INTO contests (name, date, status, judge_password, num_groups, num_judges, num_players, created_at)
                    VALUES (?, ?, 0, ?, ?, ?, ?, ?)
                ''', (name, datetime.now().isoformat(), '123456', num_groups, num_judges, num_players, datetime.now().isoformat()))
                contest_id = cursor.lastrowid
            
            # 创建默认组别
            for g in range(1, num_groups + 1):
                if get_db_type() == 'postgres':
                    cursor.execute('INSERT INTO groups (contest_id, name) VALUES (%s, %s)', (contest_id, f'第{g}组'))
                else:
                    cursor.execute('INSERT INTO groups (contest_id, name) VALUES (?, ?)', (contest_id, f'第{g}组'))
            
            # 创建默认评委
            for j in range(1, num_judges + 1):
                if get_db_type() == 'postgres':
                    cursor.execute('INSERT INTO judges (contest_id, name) VALUES (%s, %s)', (contest_id, f'评委{j}'))
                else:
                    cursor.execute('INSERT INTO judges (contest_id, name) VALUES (?, ?)', (contest_id, f'评委{j}'))
            
            conn.commit()
            conn.close()
            
            # 跳转到比赛详情页来设置选手信息
            return redirect(url_for('contest_detail', contest_id=contest_id))
        
        return render_template('create.html')
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return f"Error: {str(e)}\n\nTraceback:\n{error_trace}", 500

@app.route('/contest/<int:contest_id>')
def contest_detail(contest_id):
    try:
        conn = get_db()
        if get_db_type() == 'postgres':
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        else:
            cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contests WHERE id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contests WHERE id = ?', (contest_id,))
        
        contest = cursor.fetchone()
        
        if contest is None:
            conn.close()
            return "比赛不存在", 404
        
        # 处理POST请求 - 保存名称设置
        if request.method == 'POST':
            name_type = request.form.get('name_type')
            
            if name_type == 'group':
                # 保存组别名称
                for g in range(1, contest['num_groups'] + 1):
                    group_name = request.form.get(f'group_{g}', f'第{g}组')
                    if get_db_type() == 'postgres':
                        cursor.execute('UPDATE groups SET name = %s WHERE contest_id = %s AND id = %s', 
                                     (group_name, contest_id, g))
                    else:
                        cursor.execute('UPDATE groups SET name = ? WHERE contest_id = ? AND id = ?', 
                                     (group_name, contest_id, g))
            
            elif name_type == 'judge':
                # 保存评委姓名
                for j in range(1, contest['num_judges'] + 1):
                    judge_name = request.form.get(f'judge_{j}', f'评委{j}')
                    if get_db_type() == 'postgres':
                        cursor.execute('UPDATE judges SET name = %s WHERE contest_id = %s AND id = %s', 
                                     (judge_name, contest_id, j))
                    else:
                        cursor.execute('UPDATE judges SET name = ? WHERE contest_id = ? AND id = ?', 
                                     (judge_name, contest_id, j))
            
            elif name_type == 'player':
                # 保存选手姓名 - 需要先获取该组别的选手
                num_players = contest['num_players']
                num_groups = contest['num_groups']
                
                # 删除现有选手
                if get_db_type() == 'postgres':
                    cursor.execute('DELETE FROM contestants WHERE contest_id = %s', (contest_id,))
                else:
                    cursor.execute('DELETE FROM contestants WHERE contest_id = ?', (contest_id,))
                
                # 重新创建选手
                for g in range(1, num_groups + 1):
                    for p in range(1, num_players + 1):
                        player_name = request.form.get(f'player_{g}_{p}', f'选手{p}')
                        if get_db_type() == 'postgres':
                            cursor.execute('INSERT INTO contestants (contest_id, name, team, number) VALUES (%s, %s, %s, %s)',
                                         (contest_id, player_name, f'第{g}组', p))
                        else:
                            cursor.execute('INSERT INTO contestants (contest_id, name, team, number) VALUES (?, ?, ?, ?)',
                                         (contest_id, player_name, f'第{g}组', p))
            
            conn.commit()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contestants WHERE contest_id = %s ORDER BY number', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contestants WHERE contest_id = ? ORDER BY number', (contest_id,))
        
        contestants = cursor.fetchall()
        
        # 获取组别信息
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM groups WHERE contest_id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM groups WHERE contest_id = ?', (contest_id,))
        groups_rows = cursor.fetchall()
        groups = {i+1: row['name'] for i, row in enumerate(groups_rows)}
        
        # 获取评委信息
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM judges WHERE contest_id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM judges WHERE contest_id = ?', (contest_id,))
        judges_rows = cursor.fetchall()
        judges = {i+1: row['name'] for i, row in enumerate(judges_rows)}
        
        conn.close()
        return render_template('setup.html', comp=contest, competitions=contestants, groups=groups, judges=judges)
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/judge/login/<int:contest_id>', methods=['GET', 'POST'])
def judge_login(contest_id):
    try:
        conn = get_db()
        if get_db_type() == 'postgres':
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        else:
            cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contests WHERE id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contests WHERE id = ?', (contest_id,))
        
        contest = cursor.fetchone()
        
        # 获取组别信息
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM groups WHERE contest_id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM groups WHERE contest_id = ?', (contest_id,))
        groups_rows = cursor.fetchall()
        groups = {i+1: row['name'] for i, row in enumerate(groups_rows)}
        
        # 获取评委信息
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM judges WHERE contest_id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM judges WHERE contest_id = ?', (contest_id,))
        judges_rows = cursor.fetchall()
        judges = {i+1: row['name'] for i, row in enumerate(judges_rows)}
        
        conn.close()
        
        if request.method == 'POST':
            password = request.form['password']
            
            if contest and contest['judge_password'] == password:
                session['judge_contest_id'] = contest_id
                session['judge_logged_in'] = True
                session['judge_group'] = request.form.get('group_num', '1')
                session['judge_num'] = request.form.get('judge_num', '1')
                return redirect(url_for('scoring', contest_id=contest_id))
            
            return render_template('judge_login.html', comp=contest, groups=groups, judges=judges, error='密码错误')
        
        return render_template('judge_login.html', comp=contest, groups=groups, judges=judges, error=None)
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/scoring/<int:contest_id>', methods=['GET', 'POST'])
def scoring(contest_id):
    try:
        if not session.get('judge_logged_in') or session.get('judge_contest_id') != contest_id:
            return redirect(url_for('judge_login', contest_id=contest_id))
        
        conn = get_db()
        if get_db_type() == 'postgres':
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        else:
            cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contests WHERE id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contests WHERE id = ?', (contest_id,))
        
        contest = cursor.fetchone()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contestants WHERE contest_id = %s ORDER BY number', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contestants WHERE contest_id = ? ORDER BY number', (contest_id,))
        
        contestants = cursor.fetchall()
        
        # 构建players字典
        players = {str(p['number']): p['name'] for p in contestants}
        
        # 评分项名称（默认5个评分项）
        items = {str(i): f'评分项{i}' for i in range(1, 6)}
        
        judge_group = session.get('judge_group', '1')
        judge_num = session.get('judge_num', '1')
        
        # 获取组别和评委名称
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM groups WHERE contest_id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM groups WHERE contest_id = ?', (contest_id,))
        groups_rows = cursor.fetchall()
        group_name = groups_rows[int(judge_group)-1]['name'] if len(groups_rows) >= int(judge_group) else f'第{judge_group}组'
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM judges WHERE contest_id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM judges WHERE contest_id = ?', (contest_id,))
        judges_rows = cursor.fetchall()
        judge_name = judges_rows[int(judge_num)-1]['name'] if len(judges_rows) >= int(judge_num) else f'评委{judge_num}'
        
        if request.method == 'POST':
            selected_player = request.form.get('selected_player', '')
            
            # 找到选中的选手
            selected_contestant = None
            for c in contestants:
                if str(c['number']) == selected_player:
                    selected_contestant = c
                    break
            
            if selected_contestant:
                score1 = float(request.form.get('score_1', 0))
                score2 = float(request.form.get('score_2', 0))
                score3 = float(request.form.get('score_3', 0))
                score4 = float(request.form.get('score_4', 0))
                score5 = float(request.form.get('score_5', 0))
                total = (score1 + score2 + score3 + score4 + score5) / 5
                
                if get_db_type() == 'postgres':
                    cursor.execute('''
                        INSERT INTO scores (contest_id, contestant_id, judge_name, score1, score2, score3, score4, score5, total_score, created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ''', (contest_id, selected_contestant['id'], judge_name, score1, score2, score3, score4, score5, total, datetime.now().isoformat()))
                else:
                    cursor.execute('''
                        INSERT INTO scores (contest_id, contestant_id, judge_name, score1, score2, score3, score4, score5, total_score, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (contest_id, selected_contestant['id'], judge_name, score1, score2, score3, score4, score5, total, datetime.now().isoformat()))
                
                conn.commit()
            
            conn.close()
            session.pop('judge_logged_in', None)
            session.pop('judge_contest_id', None)
            session.pop('judge_group', None)
            session.pop('judge_num', None)
            return redirect(url_for('index'))
        
        conn.close()
        return render_template('scoring.html', comp=contest, players=players, items=items, group_name=group_name, judge_name=judge_name)
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/statistics/<int:contest_id>')
def statistics(contest_id):
    try:
        conn = get_db()
        if get_db_type() == 'postgres':
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        else:
            cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contests WHERE id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contests WHERE id = ?', (contest_id,))
        
        contest = cursor.fetchone()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contestants WHERE contest_id = %s ORDER BY number', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contestants WHERE contest_id = ? ORDER BY number', (contest_id,))
        
        contestants = cursor.fetchall()
        
        contestant_stats = []
        for contestant in contestants:
            if get_db_type() == 'postgres':
                cursor.execute('SELECT * FROM scores WHERE contestant_id = %s', (contestant['id'],))
            else:
                cursor.execute('SELECT * FROM scores WHERE contestant_id = ?', (contestant['id'],))
            
            scores = cursor.fetchall()
            
            if scores:
                avg_score = sum(s['total_score'] for s in scores) / len(scores)
                max_score = max(s['total_score'] for s in scores)
                min_score = min(s['total_score'] for s in scores)
                scores_count = len(scores)
            else:
                avg_score = 0
                max_score = 0
                min_score = 0
                scores_count = 0
            
            contestant_stats.append({
                'id': contestant['id'],
                'name': contestant['name'],
                'team': contestant['team'],
                'number': contestant['number'],
                'avg_score': avg_score,
                'max_score': max_score,
                'min_score': min_score,
                'scores_count': scores_count
            })
        
        contestant_stats.sort(key=lambda x: x['avg_score'], reverse=True)
        
        conn.close()
        return render_template('statistics.html', comp=contest, contestants=contestant_stats)
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/export/<int:contest_id>')
def export_excel(contest_id):
    try:
        conn = get_db()
        if get_db_type() == 'postgres':
            cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        else:
            cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contests WHERE id = %s', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contests WHERE id = ?', (contest_id,))
        
        contest = cursor.fetchone()
        
        if get_db_type() == 'postgres':
            cursor.execute('SELECT * FROM contestants WHERE contest_id = %s ORDER BY number', (contest_id,))
        else:
            cursor.execute('SELECT * FROM contestants WHERE contest_id = ? ORDER BY number', (contest_id,))
        
        contestants = cursor.fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "评分结果"
        
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_alignment = Alignment(horizontal='center')
        
        headers = ['序号', '选手姓名', '团队', '平均得分', '最高得分', '最低得分', '评分次数']
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        for i, contestant in enumerate(contestants, 1):
            if get_db_type() == 'postgres':
                cursor.execute('SELECT * FROM scores WHERE contestant_id = %s', (contestant['id'],))
            else:
                cursor.execute('SELECT * FROM scores WHERE contestant_id = ?', (contestant['id'],))
            
            scores = cursor.fetchall()
            
            if scores:
                avg_score = sum(s['total_score'] for s in scores) / len(scores)
                max_score = max(s['total_score'] for s in scores)
                min_score = min(s['total_score'] for s in scores)
                scores_count = len(scores)
            else:
                avg_score = 0
                max_score = 0
                min_score = 0
                scores_count = 0
            
            ws.append([contestant['number'], contestant['name'], contestant['team'], 
                       round(avg_score, 2), round(max_score, 2), round(min_score, 2), scores_count])
            
            for col in range(1, 8):
                cell = ws.cell(row=i+1, column=col)
                cell.border = thin_border
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_path)
        conn.close()
        
        return send_file(temp_path, as_attachment=True, download_name=f'{contest["name"]}_评分结果.xlsx')
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/delete/<int:contest_id>')
def delete_contest(contest_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('DELETE FROM scores WHERE contest_id = %s', (contest_id,))
            cursor.execute('DELETE FROM contestants WHERE contest_id = %s', (contest_id,))
            cursor.execute('DELETE FROM judges WHERE contest_id = %s', (contest_id,))
            cursor.execute('DELETE FROM contests WHERE id = %s', (contest_id,))
        else:
            cursor.execute('DELETE FROM scores WHERE contest_id = ?', (contest_id,))
            cursor.execute('DELETE FROM contestants WHERE contest_id = ?', (contest_id,))
            cursor.execute('DELETE FROM judges WHERE contest_id = ?', (contest_id,))
            cursor.execute('DELETE FROM contests WHERE id = ?', (contest_id,))
        
        conn.commit()
        conn.close()
        
        return redirect(url_for('index'))
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/start/<int:contest_id>')
def start_contest(contest_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('UPDATE contests SET status = 1 WHERE id = %s', (contest_id,))
        else:
            cursor.execute('UPDATE contests SET status = 1 WHERE id = ?', (contest_id,))
        
        conn.commit()
        conn.close()
        return redirect(url_for('index'))
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/end/<int:contest_id>')
def end_contest(contest_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        if get_db_type() == 'postgres':
            cursor.execute('UPDATE contests SET status = 2 WHERE id = %s', (contest_id,))
        else:
            cursor.execute('UPDATE contests SET status = 2 WHERE id = ?', (contest_id,))
        
        conn.commit()
        conn.close()
        return redirect(url_for('index'))
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/favicon.ico')
def favicon():
    return '', 204

@app.route('/favicon.png')
def favicon_png():
    return '', 204

@app.errorhandler(400)
def bad_request(error):
    return f"400 Bad Request: {str(error)}\n\nRequest Method: {request.method}\nRequest URL: {request.url}\nRequest Headers: {dict(request.headers)}\nRequest Form: {dict(request.form)}", 400

@app.errorhandler(Exception)
def handle_exception(error):
    import traceback
    error_trace = traceback.format_exc()
    return f"Error: {str(error)}\n\nRequest Method: {request.method}\nRequest URL: {request.url}\n\nTraceback:\n{error_trace}", 500

handler = app
